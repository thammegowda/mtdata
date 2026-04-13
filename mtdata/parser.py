#!/usr/bin/env python
#
# Author: Thamme Gowda [tg (at) isi (dot) edu] 
# Created: 4/4/20

import csv
from typing import Optional, Union, Tuple, List
from dataclasses import dataclass
from pathlib import Path
from mtdata import log, pbar_man, Defaults
from mtdata.entry import Entry
from itertools import zip_longest

from mtdata.utils import IO

COMPRESS_EXT = ['gz', 'bz2', 'xz']
HF_EXT = 'hfds'  # huggingface dataset
WMT21XML = 'wmt21xml'


def detect_extension(name: Union[str, Path]):
    """
    detects file extension from file name
    :param filename: name or full path of file
    :return: extension of file
    """

    filename = name if isinstance(name, str) else name.name
    parts = filename.split('.')
    ext = parts[-1]
    if ext in COMPRESS_EXT:
        ext = '.'.join(parts[-2:])
    return ext


@dataclass
class Parser:
    paths: Union[Path, List[Path]]
    ext: Optional[str] = None
    ent: Optional['Entry'] = None

    def __post_init__(self):
        if not isinstance(self.paths, list):
            self.paths = [self.paths]
        for p in self.paths:
            if isinstance(p, Path):
                assert p.exists(), f'{p} not exists'
            # skip cheks on HF datasets

        if not self.ext:
            exts = [detect_extension(p.name) for p in self.paths]
            if len(exts) == 2:
                did = self.ent and self.ent.did or ''
                log.warning(f"{did} :: Treating {' .'.join(exts)} as plain text. To override: in_ext=<ext>")
                exts = ['txt']  # treat that as plain text
            assert len(set(exts)) == 1, f'Expected a type of exts, but found: {exts}'
            self.ext = exts[0]
        assert 1 <= len(self.paths)
        # tsv and tmx just concatenate all of them
        assert len(self.paths) <= 3 or self.ext == 'tmx' or self.ext == 'tsv'

    def read_segs(self, show_pbar=True):
        readers = []
        if self.ext == 'opus_xces':
            preprocessing = 'xml'
            if "/raw/" in self.ent.in_paths[0]:
                preprocessing = 'raw'
            align, lang1_dir, lang2_dir = self.paths
            from mtdata.opus_xces import OpusXcesParser
            reader = OpusXcesParser.read(align, lang1_dir, lang2_dir, preprocessing=preprocessing)
            readers.append(reader)
        else:
            meta_fields = self.ent.meta and self.ent.meta.get('fields') or None
            for p in self.paths:
                if 'tsv' in self.ext:
                    cols = (0, 1) #extract first two columns
                    if self.ent and self.ent.cols:
                        cols = self.ent.cols
                    readers.append(self.read_tsv(p, cols=cols, meta_fields=meta_fields))
                elif 'csvwithheader' in self.ext:
                    readers.append(self.read_csv(p, meta_fields=meta_fields))
                elif 'raw' in self.ext or 'txt' in self.ext:
                    readers.append(self.read_plain(p))
                elif 'tmx' in self.ext:
                    from mtdata.tmx import read_tmx
                    readers.append(read_tmx(path=p, langs=self.ent.did.langs))
                elif 'sgm' in self.ext:
                    from mtdata.sgm import read_sgm
                    readers.append(read_sgm(p))
                elif WMT21XML in self.ext:
                    from mtdata.sgm import read_wmt21_xml
                    readers.append(read_wmt21_xml(p))
                elif HF_EXT in self.ext:
                    readers.append(self.read_hfds(p))
                elif 'xlsx' in self.ext:
                    readers.append(self.read_xlsx(p))
                else:
                    raise Exception(f'Not supported {self.ext} : {p}')

        if len(readers) == 1:
            data = readers[0]
        elif self.ext == 'tmx' or (self.ext == 'tsv' and len(self.paths) == 1):
            data = (rec for reader in readers for rec in reader)  # flatten readers from single source
        elif len(readers) == 2:
            def _zip_n_check():
                for row in zip_longest(*readers):
                    seg1, seg2 = row[:2]
                    if seg1 is None or seg2 is None:
                        raise Exception(f'{self.paths} have unequal number of segments')
                    yield seg1, seg2
            data = _zip_n_check()
        else:
            raise Exception("This is an error")
        if not show_pbar:
            yield from data
            return
        with pbar_man.counter(unit='line', desc=f"Reading {self.ent.did}") as pbar:
            for rec in data:
                yield rec
                pbar.update()

    def read_plain(self, path):
        try:
            with IO.reader(path) as stream:
                for line in stream:
                    yield line.rstrip('\r\n')
        except:
            log.warning(f'Error reading file {path}')
            raise

    def read_tsv(self, path, delim='\t', cols=None, skipheader=False, meta_fields=None):
        """
        Read data from TSV file
        :param path: path to TSV file
        :param delim: delimiter default is \\t
        :param cols: if certain columns are to be extracted;
            default is None, which returns all columns
        :return:
        """
        with IO.reader(path) as stream:
            if skipheader:
                line = stream.readline()
            for line in stream:
                row = [x.strip() for x in line.rstrip('\r\n').split(delim)]
                out_row = row
                if cols:
                    out_row = [row[idx] for idx in cols]
                    if len(cols) == 1:
                        out_row = out_row[0]  # unwrap single-column to scalar
                if meta_fields:
                    metadata = {}
                    for key, idx in meta_fields.items():
                        if key in ("source", "target") or idx >= len(row) or row[idx] in ("", None):
                            continue
                        metadata[key] = row[idx]
                    if metadata:
                        out_row.append(metadata)
                yield out_row

    def read_csv(self, path, cols=None, meta_fields=None):
        """Read data from a CSV file with header using Python's csv module.
        Handles quoted fields with embedded commas/newlines correctly.
        :param path: path to CSV file
        :param cols: column indices to extract; default is (0, 1)
        """
        if cols is None:
            cols = self.ent.cols if (self.ent and self.ent.cols) else (0, 1)
        with IO.reader(path) as stream:
            reader = csv.reader(stream)
            header = next(reader, None)  # skip header
            for row in reader:
                if not row or len(row) <= max(cols):
                    continue
                out_row = [row[c].strip() for c in cols]
                if meta_fields:
                    metadata = {}
                    for key, idx in meta_fields.items():
                        if key in ("source", "target") or idx >= len(row) or row[idx] in ("", None):
                            continue
                        metadata[key] = row[idx]
                    if metadata:
                        out_row.append(metadata)
                yield out_row

    def read_xlsx(self, path, cols=None):
        """Read data from an Excel .xlsx file.
        :param path: path to .xlsx file
        :param cols: column indices to extract; default uses ent.cols or (0, 1)
        """
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ImportError("openpyxl is required to read .xlsx files. Run: pip install openpyxl") from e
        if cols is None:
            cols = self.ent.cols if (self.ent and self.ent.cols) else (0, 1)
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
            out = [str(row[c]).strip() if row[c] is not None else '' for c in cols]
            if all(v == '' for v in out):
                continue
            yield out
        wb.close()

    @staticmethod
    def _nested_get(row, field):
        """Get a value from a dict using dot-separated path for nested access.
        e.g. _nested_get(row, "translation.ita") == row["translation"]["ita"]
        """
        parts = field.split('.')
        val = row
        for part in parts:
            val = val[part]
        return val

    def read_hfds(self, ds):
        """ Read data from huggingface Dataset
        :param ds: huggingface dataset
        :return: generator of segments
        """
        fields = self.ent.meta["fields"]  # expects dictionary
        src_field = fields['source']
        tgt_field = fields.get('target', None)
        rev_map = {v: k for k, v in fields.items()}
        # fields map is a forward map of "dest: orig", and meant to pick a subset of fields from the row
        # in the current version, I am going to retain all fields to see what all fields exist,
        # and map the subset of fields as per the dict; so, created rev_map.get(orig,orig)
        for row in ds:
            src_val = self._nested_get(row, src_field)
            tgt_val = self._nested_get(row, tgt_field) if tgt_field else None
            top_keys = {f.split('.')[0] for f in [src_field] + ([tgt_field] if tgt_field else [])}
            metadata = {rev_map.get(k, k): v for k, v in row.items() if k not in top_keys}

            src_is_list = isinstance(src_val, list)
            tgt_is_list = isinstance(tgt_val, list)
            if src_is_list and tgt_is_list:
                # Both lists (e.g. SmolDoc srcs/trgs): zip and yield each pair
                for s, t in zip(src_val, tgt_val):
                    yield [s, t, metadata]
            elif not src_is_list and tgt_is_list:
                # Source is scalar, target is list (e.g. GATITOS src/trgs): expand
                for t in tgt_val:
                    yield [src_val, t, metadata]
            else:
                out_row = [src_val]
                if tgt_val is not None:
                    out_row.append(tgt_val)
                out_row.append(metadata)
                yield out_row
